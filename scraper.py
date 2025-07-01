import requests
from bs4 import BeautifulSoup
import openpyxl
import re

def extract_emails_from_text(text):
    """Extrait tous les e-mails d'un texte"""
    email_pattern = r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b'
    return re.findall(email_pattern, text)

def get_official_website_from_parcoursup(soup):
    """Extrait le site officiel de l'établissement depuis la page Parcoursup"""
    try:
        # Chercher le lien du site officiel dans les informations de l'établissement
        website_link = soup.find('a', href=True)
        if website_link and 'http' in website_link['href']:
            return website_link['href']
    except:
        pass
    return None

def scrape_official_website(url):
    """Scrape le site officiel pour trouver des e-mails de contact"""
    try:
        response = requests.get(url, timeout=10)
        response.raise_for_status()
        soup = BeautifulSoup(response.text, 'html.parser')
        
        # Extraire tout le texte de la page et chercher des e-mails
        page_text = soup.get_text()
        emails = extract_emails_from_text(page_text)
        
        return emails[:3] if emails else []  # Retourne les 3 premiers e-mails trouvés
    except:
        return []

def scrape_and_update_excel(url, excel_file, sheet_name):
    try:
        print(f"Traitement du lien : {url}")
        response = requests.get(url, timeout=10)
        response.raise_for_status()
        soup = BeautifulSoup(response.text, 'html.parser')
        
        # Chercher la section "Contacter et échanger avec l'établissement"
        contact_section = soup.find('h3', string=lambda text: text and 'Contacter et échanger' in text)
        
        if not contact_section:
            # Essayer une recherche plus large
            contact_section = soup.find(string=lambda text: text and 'Contacter et échanger' in text)
        
        pedagogical_email = 'Non trouvé'
        admin_email = 'Non trouvé'
        general_contact = 'Non trouvé'
        
        if contact_section:
            # Trouver le conteneur parent de la section de contact
            parent_section = contact_section.find_parent() if hasattr(contact_section, 'find_parent') else contact_section.parent
            
            # Chercher dans plusieurs niveaux de parents pour trouver la section complète
            for _ in range(5):  # Essayer jusqu'à 5 niveaux de parents
                if parent_section:
                    section_text = parent_section.get_text() if hasattr(parent_section, 'get_text') else str(parent_section)
                    emails = extract_emails_from_text(section_text)
                    
                    if emails:
                        # Assigner les e-mails en fonction de leur contexte
                        for email in emails:
                            email_context = section_text.lower()
                            if 'pédagogique' in email_context or 'pedagogique' in email_context:
                                pedagogical_email = email
                            elif 'administratif' in email_context or 'administration' in email_context:
                                admin_email = email
                            elif general_contact == 'Non trouvé':
                                general_contact = email
                        
                        # Si on a trouvé des e-mails, assigner intelligemment
                        if len(emails) >= 2:
                            if pedagogical_email == 'Non trouvé':
                                pedagogical_email = emails[0]
                            if admin_email == 'Non trouvé':
                                admin_email = emails[1] if len(emails) > 1 else emails[0]
                        elif len(emails) == 1:
                            if pedagogical_email == 'Non trouvé':
                                pedagogical_email = emails[0]
                        break
                    
                    parent_section = parent_section.find_parent() if hasattr(parent_section, 'find_parent') else getattr(parent_section, 'parent', None)
                else:
                    break
        
        # Si aucun e-mail trouvé, essayer de chercher le site officiel
        if pedagogical_email == 'Non trouvé' and admin_email == 'Non trouvé':
            print("Aucun e-mail trouvé sur Parcoursup, recherche du site officiel...")
            
            # Chercher le site officiel dans la page
            official_site = get_official_website_from_parcoursup(soup)
            if official_site:
                print(f"Site officiel trouvé : {official_site}")
                emails_from_site = scrape_official_website(official_site)
                if emails_from_site:
                    pedagogical_email = emails_from_site[0] if len(emails_from_site) > 0 else 'Non trouvé'
                    admin_email = emails_from_site[1] if len(emails_from_site) > 1 else pedagogical_email
                    print(f"E-mails trouvés sur le site officiel : {emails_from_site}")
        
        # Ouvrir et mettre à jour le fichier Excel
        try:
            wb = openpyxl.load_workbook(excel_file)
            sheet = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
            
            # Ajouter une ligne avec les informations extraites
            sheet.append([url, general_contact, pedagogical_email, admin_email])
            wb.save(excel_file)
            print(f'Mise à jour réussie ! Contact général: {general_contact}, Pédagogique: {pedagogical_email}, Administratif: {admin_email}')
            
        except FileNotFoundError:
            print(f"Fichier Excel {excel_file} non trouvé. Création d'un nouveau fichier...")
            wb = openpyxl.Workbook()
            sheet = wb.active
            sheet.append(['URL', 'Contact Général', 'Mail Pédagogique', 'Mail Administratif'])  # En-têtes
            sheet.append([url, general_contact, pedagogical_email, admin_email])
            wb.save(excel_file)
            print('Nouveau fichier Excel créé avec succès!')
        except PermissionError:
            # Si le fichier est ouvert ou protégé, créer un nouveau fichier
            import datetime
            timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            new_filename = f"contacts_scrapped_{timestamp}.xlsx"
            print(f"Permission refusée pour {excel_file}. Création d'un nouveau fichier : {new_filename}")
            wb = openpyxl.Workbook()
            sheet = wb.active
            sheet.append(['URL', 'Contact Général', 'Mail Pédagogique', 'Mail Administratif'])  # En-têtes
            sheet.append([url, general_contact, pedagogical_email, admin_email])
            wb.save(new_filename)
            print(f'Nouveau fichier Excel créé : {new_filename}')
            
    except requests.exceptions.RequestException as e:
        print(f'Erreur sur le lien {url}: {e}')
        # Essayer de corriger le lien (par exemple, si le code formation est incorrect)
        if 'g_ta_cod=' in url:
            print("Tentative de correction du lien...")
            # Ici on pourrait implémenter une logique pour corriger les codes de formation
        
    except Exception as e:
        print(f'Erreur générale pour {url}: {e}')

if __name__ == '__main__':
    liens = [
        'https://dossierappel.parcoursup.fr/Candidats/public/fiches/afficherFicheFormation?g_ta_cod=43102&typeBac=0&originePc=0',
        'https://dossierappel.parcoursup.fr/Candidats/public/fiches/afficherFicheFormation?g_ta_cod=43064&typeBac=0&originePc=0',
        'https://dossierappel.parcoursup.fr/Candidats/public/fiches/afficherFicheFormation?g_ta_cod=43066&typeBac=0&originePc=0'
    ]
    for lien in liens:
        scrape_and_update_excel(lien, '20250117_cartographie_for.xlsx', 'Sheet1') 