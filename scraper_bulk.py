import requests
from bs4 import BeautifulSoup
import openpyxl
import re
import time
import datetime
from urllib.parse import urlparse

def extract_emails_from_text(text):
    """Extrait tous les e-mails d'un texte"""
    email_pattern = r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b'
    return re.findall(email_pattern, text)

def get_official_website_from_parcoursup(soup):
    """Extrait le site officiel de l'établissement depuis la page Parcoursup"""
    try:
        website_links = soup.find_all('a', href=True)
        for link in website_links:
            href = link['href']
            if 'http' in href and 'parcoursup' not in href and 'gouv.fr' not in href:
                return href
    except:
        pass
    return None

def scrape_official_website(url):
    """Scrape le site officiel pour trouver des e-mails de contact"""
    try:
        response = requests.get(url, timeout=10)
        response.raise_for_status()
        soup = BeautifulSoup(response.text, 'html.parser')
        
        page_text = soup.get_text()
        emails = extract_emails_from_text(page_text)
        
        return emails[:3] if emails else []
    except:
        return []

def extract_contacts_from_url(url):
    """Extrait les contacts d'une URL Parcoursup"""
    try:
        response = requests.get(url, timeout=15)
        response.raise_for_status()
        soup = BeautifulSoup(response.text, 'html.parser')
        
        # Chercher la section "Contacter et échanger avec l'établissement"
        contact_section = soup.find('h3', string=lambda text: text and 'Contacter et échanger' in text)
        
        if not contact_section:
            contact_section = soup.find(string=lambda text: text and 'Contacter et échanger' in text)
        
        pedagogical_email = 'Non trouvé'
        admin_email = 'Non trouvé'
        general_contact = 'Non trouvé'
        
        if contact_section:
            parent_section = contact_section.find_parent() if hasattr(contact_section, 'find_parent') else contact_section.parent
            
            for _ in range(5):
                if parent_section:
                    section_text = parent_section.get_text() if hasattr(parent_section, 'get_text') else str(parent_section)
                    emails = extract_emails_from_text(section_text)
                    
                    if emails:
                        for email in emails:
                            email_context = section_text.lower()
                            if 'pédagogique' in email_context or 'pedagogique' in email_context:
                                pedagogical_email = email
                            elif 'administratif' in email_context or 'administration' in email_context:
                                admin_email = email
                            elif general_contact == 'Non trouvé':
                                general_contact = email
                        
                        if len(emails) >= 2:
                            if pedagogical_email == 'Non trouvé':
                                pedagogical_email = emails[0]
                            if admin_email == 'Non trouvé':
                                admin_email = emails[1] if len(emails) > 1 else emails[0]
                        elif len(emails) == 1:
                            if pedagogical_email == 'Non trouvé':
                                pedagogical_email = emails[0]
                        break
                    
                    parent_section = parent_section.find_parent() if hasattr(parent_section, 'find_parent') else getattr(parent_section, 'parent', None)
                else:
                    break
        
        # Si aucun e-mail trouvé, essayer le site officiel
        if pedagogical_email == 'Non trouvé' and admin_email == 'Non trouvé':
            official_site = get_official_website_from_parcoursup(soup)
            if official_site:
                emails_from_site = scrape_official_website(official_site)
                if emails_from_site:
                    pedagogical_email = emails_from_site[0] if len(emails_from_site) > 0 else 'Non trouvé'
                    admin_email = emails_from_site[1] if len(emails_from_site) > 1 else pedagogical_email
        
        return general_contact, pedagogical_email, admin_email
        
    except Exception as e:
        print(f'Erreur pour {url}: {e}')
        return 'Erreur', 'Erreur', 'Erreur'

def process_excel_bulk(input_file, url_column='D', start_row=2):
    """Traite en masse un fichier Excel avec des URLs Parcoursup"""
    
    # Créer un fichier de sortie avec timestamp
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    output_file = f"contacts_extraits_{timestamp}.xlsx"
    
    try:
        # Ouvrir le fichier d'entrée
        wb_input = openpyxl.load_workbook(input_file)
        sheet_input = wb_input.active
        
        # Créer le fichier de sortie
        wb_output = openpyxl.Workbook()
        sheet_output = wb_output.active
        
        # En-têtes
        sheet_output.append(['Ligne', 'URL', 'Contact Général', 'Mail Pédagogique', 'Mail Administratif', 'Statut', 'Timestamp'])
        
        # Compter le nombre total de lignes
        max_row = sheet_input.max_row
        print(f"Traitement de {max_row - start_row + 1} lignes à partir de la ligne {start_row}")
        
        processed_count = 0
        error_count = 0
        
        for row_num in range(start_row, max_row + 1):
            try:
                # Extraire l'URL de la colonne spécifiée
                url_cell = sheet_input[f'{url_column}{row_num}']
                url = url_cell.value if url_cell.value else ''
                
                if not url or 'parcoursup.fr' not in str(url):
                    sheet_output.append([row_num, url, 'URL invalide', 'URL invalide', 'URL invalide', 'Skipped', datetime.datetime.now()])
                    continue
                
                print(f"Ligne {row_num}/{max_row} - Traitement de {url}")
                
                # Extraire les contacts
                general, pedagogical, admin = extract_contacts_from_url(str(url))
                
                # Ajouter au fichier de sortie
                sheet_output.append([row_num, url, general, pedagogical, admin, 'Traité', datetime.datetime.now()])
                
                processed_count += 1
                
                # Sauvegarde régulière tous les 100 traitements
                if processed_count % 100 == 0:
                    wb_output.save(output_file)
                    print(f"Sauvegarde automatique - {processed_count} lignes traitées")
                
                # Pause pour éviter la surcharge des serveurs
                time.sleep(1)  # 1 seconde entre chaque requête
                
                # Pause plus longue tous les 50 appels
                if processed_count % 50 == 0:
                    print("Pause de 10 secondes...")
                    time.sleep(10)
                
            except Exception as e:
                error_count += 1
                print(f"Erreur ligne {row_num}: {e}")
                sheet_output.append([row_num, url, 'Erreur', 'Erreur', 'Erreur', f'Erreur: {e}', datetime.datetime.now()])
                continue
        
        # Sauvegarde finale
        wb_output.save(output_file)
        
        print(f"\n=== TRAITEMENT TERMINÉ ===")
        print(f"Total traité: {processed_count}")
        print(f"Erreurs: {error_count}")
        print(f"Fichier de sortie: {output_file}")
        
    except Exception as e:
        print(f"Erreur générale: {e}")

if __name__ == '__main__':
    # Configuration
    input_file = '20250117_cartographie_for.xlsx'
    url_column = 'O'  # Colonne O contient les URLs Parcoursup
    start_row = 2     # Ligne de démarrage (2 si ligne 1 = en-têtes)
    
    print("=== DÉMARRAGE DU TRAITEMENT EN MASSE ===")
    print(f"Fichier: {input_file}")
    print(f"Colonne URL: {url_column}")
    print(f"Ligne de démarrage: {start_row}")
    print("Appuyez sur Ctrl+C pour arrêter si nécessaire")
    
    process_excel_bulk(input_file, url_column, start_row) 