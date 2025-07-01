import openpyxl

def check_excel_structure(filename):
    """Examine la structure du fichier Excel pour trouver les URLs"""
    try:
        wb = openpyxl.load_workbook(filename)
        sheet = wb.active
        
        print(f"=== ANALYSE DU FICHIER {filename} ===")
        print(f"Nombre de lignes: {sheet.max_row}")
        print(f"Nombre de colonnes: {sheet.max_column}")
        print()
        
        # Afficher les en-têtes (première ligne)
        print("=== EN-TÊTES (Ligne 1) ===")
        for col in range(1, min(sheet.max_column + 1, 11)):  # Max 10 colonnes
            cell = sheet.cell(1, col)
            col_letter = openpyxl.utils.get_column_letter(col)
            print(f"Colonne {col_letter}: '{cell.value}'")
        print()
        
        # Examiner quelques lignes de données
        print("=== ÉCHANTILLON DE DONNÉES (Lignes 2-6) ===")
        for row in range(2, min(7, sheet.max_row + 1)):
            print(f"\n--- Ligne {row} ---")
            for col in range(1, min(sheet.max_column + 1, 11)):
                cell = sheet.cell(row, col)
                col_letter = openpyxl.utils.get_column_letter(col)
                value = str(cell.value)[:50] if cell.value else "VIDE"
                print(f"{col_letter}: {value}")
        
        # Chercher les colonnes contenant des URLs Parcoursup
        print("\n=== RECHERCHE D'URLS PARCOURSUP ===")
        parcoursup_columns = []
        
        for col in range(1, sheet.max_column + 1):
            col_letter = openpyxl.utils.get_column_letter(col)
            # Vérifier les premières lignes de cette colonne
            found_parcoursup = False
            for row in range(2, min(10, sheet.max_row + 1)):
                cell = sheet.cell(row, col)
                if cell.value and 'parcoursup.fr' in str(cell.value):
                    found_parcoursup = True
                    break
            
            if found_parcoursup:
                parcoursup_columns.append(col_letter)
                print(f"✅ Colonne {col_letter} contient des URLs Parcoursup")
        
        if not parcoursup_columns:
            print("❌ Aucune URL Parcoursup trouvée dans les 10 premières lignes")
            print("Vérifiez manuellement votre fichier ou ajustez la recherche")
        else:
            print(f"\n🎯 Colonnes recommandées pour le script: {', '.join(parcoursup_columns)}")
            
    except Exception as e:
        print(f"Erreur: {e}")

if __name__ == '__main__':
    check_excel_structure('20250117_cartographie_for.xlsx') 