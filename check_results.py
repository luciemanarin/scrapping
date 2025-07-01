import openpyxl
import glob

def check_extraction_results():
    """Vérifie les résultats du scraping d'e-mails"""
    
    # Trouver le fichier de résultats le plus récent
    result_files = glob.glob("contacts_extraits_*.xlsx")
    if not result_files:
        print("❌ Aucun fichier de résultats trouvé")
        return
    
    latest_file = max(result_files)
    print(f"📁 Examen du fichier : {latest_file}")
    
    try:
        wb = openpyxl.load_workbook(latest_file)
        sheet = wb.active
        
        print(f"📊 Nombre de lignes dans le fichier : {sheet.max_row}")
        print()
        
        # Analyser les résultats
        total_rows = sheet.max_row - 1  # Exclure l'en-tête
        emails_found = 0
        errors = 0
        skipped = 0
        
        print("=== ÉCHANTILLON DES RÉSULTATS ===")
        for row in range(2, min(12, sheet.max_row + 1)):  # Afficher les 10 premières lignes
            ligne = sheet.cell(row, 1).value
            url = sheet.cell(row, 2).value
            contact_general = sheet.cell(row, 3).value
            mail_pedago = sheet.cell(row, 4).value
            mail_admin = sheet.cell(row, 5).value
            statut = sheet.cell(row, 6).value
            
            print(f"\n--- Ligne {ligne} ---")
            print(f"URL : {str(url)[:60]}...")
            print(f"Contact général : {contact_general}")
            print(f"Mail pédagogique : {mail_pedago}")
            print(f"Mail administratif : {mail_admin}")
            print(f"Statut : {statut}")
            
            # Compter les statistiques
            if statut == 'Traité':
                if mail_pedago != 'Non trouvé' or mail_admin != 'Non trouvé':
                    emails_found += 1
            elif 'Erreur' in str(statut):
                errors += 1
            elif statut == 'Skipped':
                skipped += 1
        
        print(f"\n=== STATISTIQUES GLOBALES ===")
        print(f"📈 Total de lignes traitées : {total_rows}")
        print(f"✅ Lignes avec e-mails trouvés : {emails_found}")
        print(f"❌ Lignes avec erreurs : {errors}")
        print(f"⏭️ Lignes ignorées : {skipped}")
        print(f"📧 Taux de succès e-mails : {(emails_found/total_rows*100):.1f}%" if total_rows > 0 else "N/A")
        
        # Vérifier s'il y a des e-mails valides
        if emails_found > 0:
            print(f"\n🎉 SUCCÈS ! Le script extrait bien les e-mails !")
        else:
            print(f"\n⚠️ Aucun e-mail trouvé pour l'instant. Le script continue...")
            
    except Exception as e:
        print(f"Erreur lors de la lecture : {e}")

if __name__ == '__main__':
    check_extraction_results() 